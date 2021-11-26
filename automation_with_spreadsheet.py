# Exercise 1 - Company and their products no.
# Exercise 2 - Total inventory value
# Exercise 3 - Inventory Less then 10
# Exercise 4 - find total inventory value per column and add it to new column
import openpyxl

# Reading Inventory file
inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):

    # Getting product no, supplier name, inventory and price from inventory file
    product_no = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory_price = product_list.cell(product_row, 5)

    # Exercise 1: Company and their product
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1

    # Exercise 2: Calculating Total Inventory Value
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Exercise 3: Inventory less than 10
    if inventory < 10:
        # product_under_10_inv[int(product_no)] = int(inventory)
        product_under_10_inv[supplier_name] = int(inventory)

    # Exercise 4: Adding total inventory value per column and insert to new column
    inventory_price.value = inventory * price

print(" Exercise 1: Company and their products are: ")
print(products_per_supplier)
print(" Exercise 2: Total Inventory value: ")
print(total_value_per_supplier)
print(" Exercise 3: Product under 10: ")
print(product_under_10_inv)

inventory_file.save("inventory_with_total_value.xlsx")
